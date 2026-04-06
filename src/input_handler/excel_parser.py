"""Excel-Parser v2: Liest Testfälle mit separaten Transaktionszeilen."""

from decimal import Decimal, InvalidOperation
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from src.mapping.field_mapper import parse_key_value_pairs
from src.models.testcase import (
    DebtorInfo,
    ExpectedResult,
    PaymentType,
    Standard,
    TestCase,
    TransactionInput,
)

# Pflicht-Spalten (muessen im Header vorhanden sein)
REQUIRED_COLUMNS = [
    "TestcaseID",
    "Titel",
    "Ziel",
    "Erwartetes Ergebnis",
    "Debtor IBAN",
]

VALID_PAYMENT_TYPES = {pt.value for pt in PaymentType}
VALID_EXPECTED_RESULTS = {er.value for er in ExpectedResult}


# Spalten, die eindeutig auf pacs.008-Format hindeuten.
_PACS008_MARKER_COLUMNS = {
    "InstgAgt BIC",
    "InstdAgt BIC",
    "IntrBkSttlmDt",
    "IntrBkSttlmAmt",
    "SttlmMtd",
    "BAH From BIC",
}

# Spalten, die eindeutig pain.001 sind. "Debtor IBAN" ist kein
# Marker, weil pacs.008 den Debtor-Kontoinhaber ebenfalls ueber
# IBAN identifiziert. "Zahlungstyp" mit seinen Werten SEPA/
# Domestic-QR/Domestic-IBAN/CBPR+ ist pain.001-spezifisch.
_PAIN001_MARKER_COLUMNS = {
    "Zahlungstyp",
}


def detect_message_type(header: List[str]) -> str:
    """Erkennt den Message-Type aus dem Excel-Header.

    Returns:
        "pacs.008" wenn charakteristische pacs.008-Spalten vorhanden sind,
        "pain.001" wenn pain.001-Spalten dominieren.
        Bei Ambiguitaet (beide vorhanden) wird ein ValueError geworfen.
        Wenn gar nichts passt, ebenfalls ValueError.
    """
    header_set = {h for h in header if h}
    pacs_hits = len(_PACS008_MARKER_COLUMNS & header_set)
    pain_hits = len(_PAIN001_MARKER_COLUMNS & header_set)

    if pacs_hits >= 2 and pain_hits == 0:
        return "pacs.008"
    if pain_hits >= 1 and pacs_hits == 0:
        return "pain.001"
    if pacs_hits >= 2 and pain_hits >= 1:
        raise ValueError(
            "Excel-Header enthaelt sowohl pain.001- als auch pacs.008-Spalten. "
            "Bitte verwende --message pain.001|pacs.008 um explizit zu waehlen."
        )
    raise ValueError(
        "Message-Type konnte nicht erkannt werden. Erforderlich sind entweder "
        f"mindestens eine pain.001-Spalte ({', '.join(sorted(_PAIN001_MARKER_COLUMNS))}) "
        f"oder mindestens zwei pacs.008-Spalten ({', '.join(sorted(_PACS008_MARKER_COLUMNS))})."
    )


def detect_message_type_from_file(file_path: str) -> str:
    """Liest den Header eines Excel-Files und erkennt den Message-Type."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()
    header = [str(c).strip() if c else "" for c in header_row]
    return detect_message_type(header)


def _parse_amount(raw, testcase_id: str = "", errors: list = None) -> Optional[Decimal]:
    """Parst einen Betrag, gibt None bei leerem Input zurück.

    Negative/Null-Betraege erzeugen eine Warnung statt stiller None-Rueckgabe.
    """
    if raw is None or str(raw).strip() == "":
        return None
    try:
        val = Decimal(str(raw))
        if val <= 0:
            if errors is not None:
                errors.append(
                    f"Testfall '{testcase_id}': Betrag '{raw}' ist nicht positiv."
                )
            return None
        return val
    except (InvalidOperation, TypeError, ValueError):
        if errors is not None:
            errors.append(
                f"Testfall '{testcase_id}': Betrag '{raw}' ist keine gültige Zahl."
            )
        return None


def _str_or_none(val) -> Optional[str]:
    """Gibt einen getrimmten String oder None zurück."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _parse_bool(val) -> bool:
    """Parst einen Boolean-Wert aus Excel (True/False/Ja/Nein/1/0)."""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "ja", "1", "yes", "wahr")


def _parse_optional_bool(val) -> Optional[bool]:
    """Parst einen optionalen Boolean-Wert. None bei leerem Input."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip()
    if not s:
        return None
    return s.lower() in ("true", "ja", "1", "yes", "wahr")


def _parse_transaction_input(row, col_index) -> TransactionInput:
    """Liest Transaktions-Daten aus einer Zeile."""

    def cell(col_name):
        idx = col_index.get(col_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    overrides_raw = _str_or_none(cell("Weitere Testdaten"))
    overrides = {}
    if overrides_raw:
        overrides = parse_key_value_pairs(overrides_raw)

    return TransactionInput(
        amount=_parse_amount(cell("Betrag")),
        currency=_str_or_none(cell("Währung")),
        creditor_name=_str_or_none(cell("Creditor Name")),
        creditor_iban=_str_or_none(cell("Creditor IBAN")),
        creditor_bic=_str_or_none(cell("Creditor BIC")),
        creditor_account_id=_str_or_none(cell("Creditor Kontonummer")),
        creditor_account_scheme=_str_or_none(cell("Creditor Kontoschema")),
        remittance_info=_str_or_none(cell("Verwendungszweck")),
        purpose_code=_str_or_none(cell("Verwendungszweck-Code")),
        overrides=overrides,
    )


def parse_excel(file_path: str) -> Tuple[List[TestCase], List[str]]:
    """Liest und validiert eine Excel-Datei mit Testfällen (v2-Format).

    Zeilen mit TestcaseID = neuer Testfall.
    Zeilen ohne TestcaseID = zusaetzliche Transaktion zum vorherigen Testfall.

    Returns:
        Tuple von (testcases, errors). Bei Fehlern ist testcases leer.
    """
    errors = []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return [], [f"Excel-Datei konnte nicht geöffnet werden: {e}"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return [], ["Die Excel-Datei ist leer."]

    # Header-Zeile validieren
    header = [str(cell).strip() if cell else "" for cell in rows[0]]

    missing_columns = [col for col in REQUIRED_COLUMNS if col not in header]
    if missing_columns:
        return [], [
            f"Fehlende Pflichtspalten: {', '.join(missing_columns)}. "
            f"Erwartet mindestens: {', '.join(REQUIRED_COLUMNS)}"
        ]

    col_index = {name: i for i, name in enumerate(header)}

    def cell_val(row, col_name):
        idx = col_index.get(col_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # Datenzeilen verarbeiten
    testcases = []
    current_tc = None
    seen_ids = set()

    for row in rows[1:]:
        testcase_id = _str_or_none(cell_val(row, "TestcaseID"))

        if testcase_id:
            # === Neue Testfall-Zeile ===
            if current_tc is not None:
                testcases.append(current_tc)

            row_errors = []

            # Duplikat-Erkennung
            if testcase_id in seen_ids:
                row_errors.append(
                    f"Testfall '{testcase_id}': TestcaseID ist doppelt vorhanden."
                )
            seen_ids.add(testcase_id)

            titel = _str_or_none(cell_val(row, "Titel"))
            if not titel:
                row_errors.append(f"Testfall '{testcase_id}': 'Titel' fehlt.")
                titel = testcase_id

            ziel = _str_or_none(cell_val(row, "Ziel"))
            if not ziel:
                row_errors.append(f"Testfall '{testcase_id}': 'Ziel' fehlt.")
                ziel = ""

            expected_raw = _str_or_none(cell_val(row, "Erwartetes Ergebnis"))
            if not expected_raw or expected_raw not in VALID_EXPECTED_RESULTS:
                row_errors.append(
                    f"Testfall '{testcase_id}': 'Erwartetes Ergebnis' muss 'OK' oder 'NOK' sein, "
                    f"ist aber '{expected_raw}'."
                )
                errors.extend(row_errors)
                current_tc = None
                continue
            expected_result = ExpectedResult(expected_raw)

            debtor_iban = _str_or_none(cell_val(row, "Debtor IBAN"))
            if not debtor_iban:
                row_errors.append(f"Testfall '{testcase_id}': 'Debtor IBAN' fehlt.")
                errors.extend(row_errors)
                current_tc = None
                continue

            if row_errors:
                errors.extend(row_errors)

            # Optionale Felder
            payment_type_raw = _str_or_none(cell_val(row, "Zahlungstyp"))
            payment_type = None
            if payment_type_raw:
                if payment_type_raw in VALID_PAYMENT_TYPES:
                    payment_type = PaymentType(payment_type_raw)
                else:
                    errors.append(
                        f"Testfall '{testcase_id}': Ungültiger Zahlungstyp '{payment_type_raw}'. "
                        f"Gültig: {', '.join(VALID_PAYMENT_TYPES)}"
                    )
                    current_tc = None
                    continue

            debtor_name = _str_or_none(cell_val(row, "Debtor Name"))
            debtor_bic = _str_or_none(cell_val(row, "Debtor BIC"))

            debtor = DebtorInfo(
                name=debtor_name,
                iban=debtor_iban,
                bic=debtor_bic,
                street=_str_or_none(cell_val(row, "Debtor Strasse")),
                building=_str_or_none(cell_val(row, "Debtor Hausnummer")),
                postal_code=_str_or_none(cell_val(row, "Debtor PLZ")),
                town=_str_or_none(cell_val(row, "Debtor Ort")),
                country=_str_or_none(cell_val(row, "Debtor Land")) or "CH",
            )

            violate_rule = _str_or_none(cell_val(row, "ViolateRule"))

            overrides_raw = _str_or_none(cell_val(row, "Weitere Testdaten"))
            overrides = {}
            group_id = None
            if overrides_raw:
                overrides = parse_key_value_pairs(overrides_raw)
                if "ViolateRule" in overrides:
                    violate_rule = violate_rule or overrides.pop("ViolateRule")
                if "GroupId" in overrides:
                    group_id = overrides.pop("GroupId")

            # Standard (optional)
            standard_raw = _str_or_none(cell_val(row, "Standard"))
            standard = Standard.SPS_2025
            if standard_raw:
                standard_lower = standard_raw.lower().strip()
                valid_standards = {s.value: s for s in Standard}
                if standard_lower in valid_standards:
                    standard = valid_standards[standard_lower]
                else:
                    errors.append(
                        f"Testfall '{testcase_id}': Ungültiger Standard '{standard_raw}'. "
                        f"Gültig: {', '.join(valid_standards.keys())}"
                    )
                    current_tc = None
                    continue

            first_tx = _parse_transaction_input(row, col_index)

            # Instant-Flag (optional)
            instant = _parse_bool(cell_val(row, "Instant"))

            # Sammelauftrag / Batch Booking (optional)
            batch_booking = _parse_optional_bool(cell_val(row, "Sammelauftrag"))

            current_tc = TestCase(
                testcase_id=testcase_id,
                titel=titel,
                ziel=ziel,
                expected_result=expected_result,
                payment_type=payment_type,
                amount=first_tx.amount,
                currency=first_tx.currency,
                debtor=debtor,
                instant=instant,
                batch_booking=batch_booking,
                overrides=overrides,
                violate_rule=violate_rule,
                transaction_inputs=[first_tx],
                standard=standard,
                group_id=group_id,
                expected_api_response=_str_or_none(cell_val(row, "Erwartete API-Antwort")) or "",
                remarks=_str_or_none(cell_val(row, "Bemerkungen")) or "",
            )

        else:
            # === Transaktionszeile (kein TestcaseID) ===
            if current_tc is None:
                continue

            tx_input = _parse_transaction_input(row, col_index)
            current_tc.transaction_inputs.append(tx_input)

    if current_tc is not None:
        testcases.append(current_tc)

    wb.close()

    if errors:
        return [], errors

    return testcases, []


# ---------------------------------------------------------------------------
# pacs.008 Excel Parser
# ---------------------------------------------------------------------------

# Pflicht-Spalten fuer pacs.008-Testfaelle.
PACS008_REQUIRED_COLUMNS = [
    "TestcaseID",
    "Titel",
    "Ziel",
    "Erwartetes Ergebnis",
]


def _pacs008_postal_address(row, col_index, prefix: str):
    """Baut eine PostalAddress aus Debtor-/Creditor-Spalten mit Prefix."""
    from src.models.pacs008 import PostalAddress

    def cell(col_name):
        idx = col_index.get(col_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    street = _str_or_none(cell(f"{prefix} Strasse"))
    bldg = _str_or_none(cell(f"{prefix} Hausnummer"))
    plz = _str_or_none(cell(f"{prefix} PLZ"))
    town = _str_or_none(cell(f"{prefix} Ort"))
    ctry = _str_or_none(cell(f"{prefix} Land"))

    if not any([street, bldg, plz, town, ctry]):
        return None

    return PostalAddress(
        street_name=street,
        building_number=bldg,
        postal_code=plz,
        town_name=town,
        country=ctry,
    )


def parse_pacs008_excel(file_path: str):
    """Parst ein pacs.008-Testfall-Excel.

    Returns:
        Tuple aus (List[Pacs008TestCase], List[str]). Bei Parser-Fehlern
        ist die Testcase-Liste leer und die Error-Liste gefuellt.

    Unterstuetzt alle Spalten aus dem Plan (WP-03 Spalten-Set). Alle
    nicht-kritischen Spalten sind optional; fehlende Spalten erzeugen
    keine Warnung. Uebersteuerbare Dot-Notation-Overrides werden aus
    'Weitere Testdaten' gelesen.
    """
    from src.models.pacs008 import Pacs008Flavor, Pacs008TestCase, SettlementMethod
    from src.models.testcase import ExpectedResult

    errors: List[str] = []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return [], [f"Excel-Datei konnte nicht geoeffnet werden: {e}"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return [], ["Die Excel-Datei ist leer."]

    header = [str(c).strip() if c else "" for c in rows[0]]
    missing = [c for c in PACS008_REQUIRED_COLUMNS if c not in header]
    if missing:
        return [], [
            f"Fehlende Pflichtspalten: {', '.join(missing)}. "
            f"Erwartet mindestens: {', '.join(PACS008_REQUIRED_COLUMNS)}"
        ]

    col_index = {name: i for i, name in enumerate(header)}

    def cell_val(row, col_name):
        idx = col_index.get(col_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    valid_expected = {er.value for er in ExpectedResult}
    valid_flavors = {f.value for f in Pacs008Flavor}
    valid_sttlm = {s.value for s in SettlementMethod}

    testcases: List = []
    seen_ids = set()

    for row_idx, row in enumerate(rows[1:], start=2):
        tc_id = _str_or_none(cell_val(row, "TestcaseID"))
        if not tc_id:
            # Continuation-Rows sind in V1 nicht unterstuetzt (Single-Tx-Fokus),
            # leere Zeilen werden ignoriert.
            continue

        row_errors = []

        if tc_id in seen_ids:
            row_errors.append(f"Testfall '{tc_id}': TestcaseID ist doppelt vorhanden.")
        seen_ids.add(tc_id)

        titel = _str_or_none(cell_val(row, "Titel"))
        if not titel:
            row_errors.append(f"Testfall '{tc_id}': 'Titel' fehlt.")
            titel = tc_id

        ziel = _str_or_none(cell_val(row, "Ziel"))
        if not ziel:
            row_errors.append(f"Testfall '{tc_id}': 'Ziel' fehlt.")
            ziel = ""

        expected_raw = _str_or_none(cell_val(row, "Erwartetes Ergebnis"))
        if not expected_raw or expected_raw not in valid_expected:
            row_errors.append(
                f"Testfall '{tc_id}': 'Erwartetes Ergebnis' muss OK/NOK sein, "
                f"ist aber '{expected_raw}'."
            )
            errors.extend(row_errors)
            continue
        expected_result = ExpectedResult(expected_raw)

        # Flavor
        flavor_raw = _str_or_none(cell_val(row, "Flavor"))
        flavor = Pacs008Flavor.CBPR_PLUS
        if flavor_raw:
            if flavor_raw in valid_flavors:
                flavor = Pacs008Flavor(flavor_raw)
            else:
                row_errors.append(
                    f"Testfall '{tc_id}': Ungueltiger Flavor '{flavor_raw}'. "
                    f"Gueltig: {', '.join(sorted(valid_flavors))}"
                )

        # Settlement Method
        sttlm_raw = _str_or_none(cell_val(row, "SttlmMtd"))
        sttlm = SettlementMethod.INDA
        if sttlm_raw:
            if sttlm_raw in valid_sttlm:
                sttlm = SettlementMethod(sttlm_raw)
            else:
                row_errors.append(
                    f"Testfall '{tc_id}': Ungueltige SttlmMtd '{sttlm_raw}'. "
                    f"Gueltig: {', '.join(sorted(valid_sttlm))}"
                )

        # Overrides
        overrides_raw = _str_or_none(cell_val(row, "Weitere Testdaten"))
        overrides = {}
        if overrides_raw:
            overrides = parse_key_value_pairs(overrides_raw)

        violate_rule = _str_or_none(cell_val(row, "ViolateRule"))
        if "ViolateRule" in overrides:
            violate_rule = violate_rule or overrides.pop("ViolateRule")

        # Addresses
        debtor_addr = _pacs008_postal_address(row, col_index, "Debtor")
        creditor_addr = _pacs008_postal_address(row, col_index, "Creditor")

        # Amount (pacs.008 nutzt 'IntrBkSttlmAmt' als Hauptbetrag; wir akzeptieren
        # auch 'Betrag' als Alias fuer Komfort)
        amount_raw = cell_val(row, "IntrBkSttlmAmt")
        if amount_raw is None:
            amount_raw = cell_val(row, "Betrag")
        amount = _parse_amount(amount_raw, tc_id, errors=row_errors) if amount_raw is not None else None

        currency = _str_or_none(cell_val(row, "Währung"))
        if not currency:
            currency = _str_or_none(cell_val(row, "Waehrung"))

        if row_errors:
            errors.extend(row_errors)
            continue

        tc = Pacs008TestCase(
            testcase_id=tc_id,
            titel=titel,
            ziel=ziel,
            expected_result=expected_result,
            flavor=flavor,
            bah_from_bic=_str_or_none(cell_val(row, "BAH From BIC")),
            bah_to_bic=_str_or_none(cell_val(row, "BAH To BIC")),
            instructing_agent_bic=_str_or_none(cell_val(row, "InstgAgt BIC")),
            instructing_agent_clr_sys_mmb_id=_str_or_none(cell_val(row, "InstgAgt ClrSysMmbId")),
            instructed_agent_bic=_str_or_none(cell_val(row, "InstdAgt BIC")),
            instructed_agent_clr_sys_mmb_id=_str_or_none(cell_val(row, "InstdAgt ClrSysMmbId")),
            settlement_method=sttlm,
            interbank_settlement_date=_str_or_none(cell_val(row, "IntrBkSttlmDt")),
            charge_bearer=_str_or_none(cell_val(row, "ChrgBr")),
            debtor_name=_str_or_none(cell_val(row, "Debtor Name")),
            debtor_address=debtor_addr,
            debtor_iban=_str_or_none(cell_val(row, "Debtor IBAN")),
            debtor_account_other_id=_str_or_none(cell_val(row, "Debtor Kontonummer")),
            debtor_account_other_scheme=_str_or_none(cell_val(row, "Debtor Kontoschema")),
            debtor_agent_bic=_str_or_none(cell_val(row, "DbtrAgt BIC")),
            debtor_agent_clr_sys_mmb_id=_str_or_none(cell_val(row, "DbtrAgt ClrSysMmbId")),
            creditor_name=_str_or_none(cell_val(row, "Creditor Name")),
            creditor_address=creditor_addr,
            creditor_iban=_str_or_none(cell_val(row, "Creditor IBAN")),
            creditor_account_other_id=_str_or_none(cell_val(row, "Creditor Kontonummer")),
            creditor_account_other_scheme=_str_or_none(cell_val(row, "Creditor Kontoschema")),
            creditor_agent_bic=_str_or_none(cell_val(row, "CdtrAgt BIC")),
            creditor_agent_clr_sys_mmb_id=_str_or_none(cell_val(row, "CdtrAgt ClrSysMmbId")),
            intermediary_agent_1_bic=_str_or_none(cell_val(row, "IntrmyAgt1 BIC")),
            intermediary_agent_1_clr_sys_mmb_id=_str_or_none(cell_val(row, "IntrmyAgt1 ClrSysMmbId")),
            intermediary_agent_2_bic=_str_or_none(cell_val(row, "IntrmyAgt2 BIC")),
            intermediary_agent_3_bic=_str_or_none(cell_val(row, "IntrmyAgt3 BIC")),
            amount=amount,
            currency=currency,
            purpose_code=_str_or_none(cell_val(row, "PurposeCode")),
            category_purpose=_str_or_none(cell_val(row, "CategoryPurpose")),
            remittance_info=_str_or_none(cell_val(row, "Verwendungszweck")),
            uetr=_str_or_none(cell_val(row, "UETR")),
            violate_rule=violate_rule,
            overrides=overrides,
            expected_api_response=_str_or_none(cell_val(row, "Erwartete API-Antwort")),
            remarks=_str_or_none(cell_val(row, "Bemerkungen")),
        )
        testcases.append(tc)

    wb.close()

    if errors:
        return [], errors

    return testcases, []
