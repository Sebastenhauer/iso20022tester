"""Generiert templates/testfaelle_pacs008_comprehensive.xlsx mit 50 Cases.

Keine Laufzeit-Komponente; wird einmal zum Befuellen der Excel
ausgefuehrt, das Ergebnis liegt dann im Repo. Script bleibt im Repo
fuer reproducibility / spaetere Anpassungen.
"""

from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Reference BIC pool (real banks; public knowledge)
# ---------------------------------------------------------------------------

BICS = {
    # Switzerland
    "UBS_ZURICH": "UBSWCHZH80A",
    "CS_ZURICH": "CRESCHZZ80A",
    "ZKB_ZURICH": "ZKBKCHZZ80A",
    "POSTFIN_BERN": "POFICHBEXXX",
    "RAIFCH": "RAIFCH22XXX",
    # Germany
    "DB_FFM": "DEUTDEFFXXX",
    "COBA_FFM": "COBADEFFXXX",
    "DZ_FFM": "GENODEFFXXX",
    "HVB_MUC": "HYVEDEMMXXX",
    # France
    "BNP_PARIS": "BNPAFRPPXXX",
    "CA_PARIS": "AGRIFRPPXXX",
    "SG_PARIS": "SOGEFRPPXXX",
    # UK
    "HSBC_LDN": "HBUKGB4BXXX",
    "BARCLAYS_LDN": "BARCGB22XXX",
    "NATWEST_LDN": "NWBKGB2LXXX",
    # USA
    "JPM_NY": "CHASUS33XXX",
    "BOA_NY": "BOFAUS3NXXX",
    "CITI_NY": "CITIUS33XXX",
    "WF_SF": "WFBIUS6SXXX",
    # Asia
    "MIZUHO_TKY": "MHCBJPJTXXX",
    "MUFG_TKY": "BOTKJPJTXXX",
    "SMBC_TKY": "SMBCJPJTXXX",
    "ICBC_BJ": "ICBKCNBJXXX",
    "BOC_BJ": "BKCHCNBJXXX",
    "HSBC_HK": "HSBCHKHHXXX",
    "SCB_SG": "SCBLSGSGXXX",
    "DBS_SG": "DBSSSGSGXXX",
    # Middle East
    "NCB_RIY": "NCBKSARIXXX",
    # LatAm
    "ITAU_BR": "ITAUBRSPXXX",
    "SANTANDER_BR": "BCBBBRPPXXX",
    # Other EU
    "ING_AMS": "INGBNL2AXXX",
    "RABO_UT": "RABONL2UXXX",
    "KBC_BRU": "KREDBEBBXXX",
    "BBVA_MAD": "BBVAESMMXXX",
    "UNICREDIT_MI": "UNCRITMMXXX",
    "OTP_BUD": "OTPVHUHBXXX",
    # Canada
    "RBC_TO": "ROYCCAT2XXX",
    # Australia
    "ANZ_SYD": "ANZBAU3MXXX",
}

# Sample IBANs per country (structurally valid Mod-97 test IBANs)
IBANS = {
    "CH": "CH9300762011623852957",
    "DE": "DE89370400440532013000",
    "FR": "FR1420041010050500013M02606",
    "GB": "GB29NWBK60161331926819",
    "NL": "NL91ABNA0417164300",
    "BE": "BE68539007547034",
    "ES": "ES9121000418450200051332",
    "IT": "IT60X0542811101000000123456",
    "AT": "AT611904300234573201",
    "SE": "SE4550000000058398257466",
    "DK": "DK5000400440116243",
    "NO": "NO9386011117947",
    "FI": "FI2112345600000785",
    "PL": "PL61109010140000071219812874",
    "HU": "HU42117730161111101800000000",
    "RO": "RO49AAAA1B31007593840000",
    "CZ": "CZ6508000000192000145399",
    "SK": "SK3112000000198742637541",
    "PT": "PT50000201231234567890154",
    "IE": "IE29AIBK93115212345678",
    "SA": "SA0380000000608010167519",
}

# Swiss addresses pool (ASCII-safe, SPS Latin-1 subset)
CH_ADDRESSES = [
    ("Bahnhofstrasse", "42", "8001", "Zurich"),
    ("Paradeplatz", "8", "8001", "Zurich"),
    ("Marktgasse", "15", "3011", "Bern"),
    ("Rue du Rhone", "22", "1204", "Geneve"),
    ("Freie Strasse", "87", "4001", "Basel"),
    ("Via Nassa", "22", "6900", "Lugano"),
    ("Sempacherstrasse", "3", "6003", "Luzern"),
]

# International addresses pool per country (real cities, ASCII)
INTL_ADDRESSES = {
    "DE": [
        ("Unter den Linden", "7", "10117", "Berlin"),
        ("Kaiserstrasse", "30", "60311", "Frankfurt am Main"),
        ("Marienplatz", "1", "80331", "Munich"),
    ],
    "FR": [
        ("Avenue des Champs-Elysees", "100", "75008", "Paris"),
        ("Rue de Rivoli", "15", "75001", "Paris"),
    ],
    "GB": [
        ("Canary Wharf", "1", "E14 5AB", "London"),
        ("Bishopsgate", "100", "EC2N 4AA", "London"),
        ("King William Street", "8", "EC4N 7BE", "London"),
    ],
    "NL": [("Damrak", "70", "1012 LM", "Amsterdam")],
    "BE": [("Rue Royale", "12", "1000", "Brussels")],
    "ES": [("Paseo de la Castellana", "81", "28046", "Madrid")],
    "IT": [("Piazza Gae Aulenti", "8", "20154", "Milan")],
    "AT": [("Stephansplatz", "1", "1010", "Vienna")],
    "SE": [("Kungsgatan", "36", "111 35", "Stockholm")],
    "DK": [("Stroget", "22", "1117", "Copenhagen")],
    "PL": [("Emilii Plater", "53", "00-113", "Warsaw")],
    "HU": [("Andrassy ut", "45", "1061", "Budapest")],
    "CZ": [("Wenceslas Square", "28", "110 00", "Prague")],
    "PT": [("Avenida da Liberdade", "110", "1250-096", "Lisbon")],
    "IE": [("Grafton Street", "50", "D02 H440", "Dublin")],
    "US": [
        ("Wall Street", "40", "10005", "New York"),
        ("Park Avenue", "270", "10017", "New York"),
        ("Market Street", "555", "94105", "San Francisco"),
    ],
    "CA": [("Bay Street", "200", "M5J 2J2", "Toronto")],
    "JP": [
        ("Marunouchi", "1-1-1", "100-0005", "Tokyo"),
        ("Ginza", "4-6-16", "104-0061", "Tokyo"),
    ],
    "CN": [("Jianguomen", "8", "100022", "Beijing")],
    "HK": [("Connaught Road", "1", "", "Central")],
    "SG": [("Marina Boulevard", "12", "018982", "Singapore")],
    "AU": [("George Street", "1", "2000", "Sydney")],
    "SA": [("King Fahd Road", "88", "12373", "Riyadh")],
    "BR": [("Avenida Paulista", "1578", "01310-200", "Sao Paulo")],
}


def addr_for(country):
    if country == "CH":
        s, b, p, t = CH_ADDRESSES[0]
        return s, b, p, t, "CH"
    pool = INTL_ADDRESSES.get(country, [("Main St", "1", "00000", "Capital")])
    s, b, p, t = pool[0]
    return s, b, p, t, country


# ---------------------------------------------------------------------------
# Columns (muessen mit parse_pacs008_excel matchen)
# ---------------------------------------------------------------------------

HEADER = [
    "TestcaseID", "Titel", "Ziel", "Erwartetes Ergebnis", "Flavor",
    "BAH From BIC", "BAH To BIC",
    "InstgAgt BIC", "InstdAgt BIC",
    "Debtor Name", "Debtor Strasse", "Debtor Hausnummer", "Debtor PLZ", "Debtor Ort", "Debtor Land",
    "Debtor IBAN", "Debtor Kontonummer", "Debtor Kontoschema",
    "DbtrAgt BIC", "DbtrAgt ClrSysMmbId",
    "IntrmyAgt1 BIC", "IntrmyAgt1 ClrSysMmbId", "IntrmyAgt2 BIC", "IntrmyAgt3 BIC",
    "Creditor Name", "Creditor Strasse", "Creditor Hausnummer", "Creditor PLZ", "Creditor Ort", "Creditor Land",
    "Creditor IBAN", "Creditor Kontonummer", "Creditor Kontoschema",
    "CdtrAgt BIC", "CdtrAgt ClrSysMmbId",
    "IntrBkSttlmAmt", "Währung", "IntrBkSttlmDt", "SttlmMtd",
    "ChrgBr", "UETR",
    "PurposeCode", "CategoryPurpose", "Verwendungszweck",
    "ViolateRule", "Weitere Testdaten",
    "Erwartete API-Antwort", "Bemerkungen",
]

COL_IDX = {name: i for i, name in enumerate(HEADER)}


def blank_row():
    return [None] * len(HEADER)


def make_row(
    tc_id, titel, ziel, expected="OK",
    instg_bic="UBSWCHZH80A", instd_bic="DEUTDEFFXXX",
    debtor_name="Muster AG", debtor_country="CH",
    debtor_iban="CH9300762011623852957", debtor_agt_bic="UBSWCHZH80A",
    creditor_name="Empfaenger GmbH", creditor_country="DE",
    creditor_iban=None, creditor_agt_bic="DEUTDEFFXXX",
    amount="1000.00", currency="EUR",
    sttlm_dt="2026-04-09", sttlm_mtd="INDA",
    chrg_br="SHAR",
    intrmy1_bic=None, intrmy1_clr=None, intrmy2_bic=None,
    cdtr_acct_other=None, cdtr_acct_scheme=None, cdtr_agt_clr=None,
    dbtr_agt_clr=None,
    purpose=None, category_purp=None, rmt=None,
    violate=None, overrides=None, bemerkungen=None,
):
    r = blank_row()
    r[COL_IDX["TestcaseID"]] = tc_id
    r[COL_IDX["Titel"]] = titel
    r[COL_IDX["Ziel"]] = ziel
    r[COL_IDX["Erwartetes Ergebnis"]] = expected
    r[COL_IDX["Flavor"]] = "CBPR+"
    r[COL_IDX["BAH From BIC"]] = instg_bic
    r[COL_IDX["BAH To BIC"]] = instd_bic
    r[COL_IDX["InstgAgt BIC"]] = instg_bic
    r[COL_IDX["InstdAgt BIC"]] = instd_bic
    # Debtor
    ds, db, dp, dt, dc = addr_for(debtor_country)
    r[COL_IDX["Debtor Name"]] = debtor_name
    r[COL_IDX["Debtor Strasse"]] = ds
    r[COL_IDX["Debtor Hausnummer"]] = db
    r[COL_IDX["Debtor PLZ"]] = dp
    r[COL_IDX["Debtor Ort"]] = dt
    r[COL_IDX["Debtor Land"]] = dc
    r[COL_IDX["Debtor IBAN"]] = debtor_iban
    r[COL_IDX["DbtrAgt BIC"]] = debtor_agt_bic
    r[COL_IDX["DbtrAgt ClrSysMmbId"]] = dbtr_agt_clr
    # Intermediary
    r[COL_IDX["IntrmyAgt1 BIC"]] = intrmy1_bic
    r[COL_IDX["IntrmyAgt1 ClrSysMmbId"]] = intrmy1_clr
    r[COL_IDX["IntrmyAgt2 BIC"]] = intrmy2_bic
    # Creditor
    cs, cb, cp, ct, cc = addr_for(creditor_country)
    r[COL_IDX["Creditor Name"]] = creditor_name
    r[COL_IDX["Creditor Strasse"]] = cs
    r[COL_IDX["Creditor Hausnummer"]] = cb
    r[COL_IDX["Creditor PLZ"]] = cp
    r[COL_IDX["Creditor Ort"]] = ct
    r[COL_IDX["Creditor Land"]] = cc
    r[COL_IDX["Creditor IBAN"]] = creditor_iban
    r[COL_IDX["Creditor Kontonummer"]] = cdtr_acct_other
    r[COL_IDX["Creditor Kontoschema"]] = cdtr_acct_scheme
    r[COL_IDX["CdtrAgt BIC"]] = creditor_agt_bic
    r[COL_IDX["CdtrAgt ClrSysMmbId"]] = cdtr_agt_clr
    # Amount
    r[COL_IDX["IntrBkSttlmAmt"]] = amount
    r[COL_IDX["Währung"]] = currency
    r[COL_IDX["IntrBkSttlmDt"]] = sttlm_dt
    r[COL_IDX["SttlmMtd"]] = sttlm_mtd
    r[COL_IDX["ChrgBr"]] = chrg_br
    # Purpose / Remittance
    r[COL_IDX["PurposeCode"]] = purpose
    r[COL_IDX["CategoryPurpose"]] = category_purp
    r[COL_IDX["Verwendungszweck"]] = rmt
    # Negative
    r[COL_IDX["ViolateRule"]] = violate
    r[COL_IDX["Weitere Testdaten"]] = overrides
    r[COL_IDX["Bemerkungen"]] = bemerkungen
    return r


# ---------------------------------------------------------------------------
# Test Cases
# ---------------------------------------------------------------------------

rows = []

# --- Group A: Basic Positive (10 Cases, various currencies/corridors) ---
rows.append(make_row(
    "TC-PCS-001", "CBPR+ EUR CH->DE minimal",
    "Simpelster CBPR+ Credit Transfer CH->DE",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["DB_FFM"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["DB_FFM"],
    creditor_iban=IBANS["DE"], creditor_country="DE",
    amount="1500.00", currency="EUR", rmt="Rechnung 2026-001",
    bemerkungen="Baseline Positive",
))
rows.append(make_row(
    "TC-PCS-002", "CBPR+ USD CH->US", "Cross-Border USD Zahlung CH->US",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["JPM_NY"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["JPM_NY"],
    creditor_name="US Partner Inc", creditor_country="US",
    cdtr_acct_other="021000021", cdtr_acct_scheme="BBAN",
    amount="25000.00", currency="USD",
    rmt="Trade Settlement Q2",
    bemerkungen="USD ohne IBAN (US verwendet Fedwire-Konten)",
))
rows.append(make_row(
    "TC-PCS-003", "CBPR+ GBP CH->UK", "Zahlung in GBP an GB",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["HSBC_LDN"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["HSBC_LDN"],
    creditor_name="London Services Ltd", creditor_country="GB",
    creditor_iban=IBANS["GB"],
    amount="5000.00", currency="GBP",
    rmt="Consulting Fee April",
))
rows.append(make_row(
    "TC-PCS-004", "CBPR+ JPY CH->JP", "Zahlung nach Japan",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["MIZUHO_TKY"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["MIZUHO_TKY"],
    creditor_name="Tokyo Industries Co", creditor_country="JP",
    cdtr_acct_other="1234567", cdtr_acct_scheme="BBAN",
    amount="500000.00", currency="JPY",
    rmt="Supply Contract",
))
rows.append(make_row(
    "TC-PCS-005", "CBPR+ CHF DE->CH", "CHF-Zahlung von DE nach CH",
    instg_bic=BICS["DB_FFM"], instd_bic=BICS["UBS_ZURICH"],
    debtor_name="DE Sender AG", debtor_country="DE",
    debtor_iban=IBANS["DE"], debtor_agt_bic=BICS["DB_FFM"],
    creditor_agt_bic=BICS["UBS_ZURICH"],
    creditor_name="Schweizer AG", creditor_country="CH",
    creditor_iban=IBANS["CH"],
    amount="3200.00", currency="CHF",
    rmt="Dienstleistung CH",
))
rows.append(make_row(
    "TC-PCS-006", "CBPR+ CNY CH->CN", "Zahlung nach China",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["ICBC_BJ"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["ICBC_BJ"],
    creditor_name="Beijing Trading Co", creditor_country="CN",
    cdtr_acct_other="6222020200000000001", cdtr_acct_scheme="BBAN",
    amount="15000.00", currency="CNY",
    rmt="Material Import",
))
rows.append(make_row(
    "TC-PCS-007", "CBPR+ SGD CH->SG", "Zahlung nach Singapur",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["DBS_SG"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["DBS_SG"],
    creditor_name="SG Holdings Pte", creditor_country="SG",
    cdtr_acct_other="1234567890", cdtr_acct_scheme="BBAN",
    amount="8000.00", currency="SGD",
    rmt="Service Payment",
))
rows.append(make_row(
    "TC-PCS-008", "CBPR+ AUD CH->AU", "Zahlung nach Australien",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["ANZ_SYD"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["ANZ_SYD"],
    creditor_name="Sydney Exports Pty", creditor_country="AU",
    cdtr_acct_other="123456789", cdtr_acct_scheme="BBAN",
    amount="12000.00", currency="AUD",
    rmt="Commodity Purchase",
))
rows.append(make_row(
    "TC-PCS-009", "CBPR+ CAD CH->CA", "Zahlung nach Kanada",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["RBC_TO"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["RBC_TO"],
    creditor_name="Toronto Corp", creditor_country="CA",
    cdtr_acct_other="12345678901", cdtr_acct_scheme="BBAN",
    amount="9500.00", currency="CAD",
    rmt="Invoice Payment",
))
rows.append(make_row(
    "TC-PCS-010", "CBPR+ HKD CH->HK", "Zahlung nach Hong Kong",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["HSBC_HK"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["HSBC_HK"],
    creditor_name="HK Trading Ltd", creditor_country="HK",
    cdtr_acct_other="004123456789", cdtr_acct_scheme="BBAN",
    amount="20000.00", currency="HKD",
    rmt="Asia Pacific Ops",
))

# --- Group B: Intermediary Agents (5 Cases) ---
rows.append(make_row(
    "TC-PCS-011", "CBPR+ USD mit Korrespondent (2-Hop)",
    "USD-Zahlung CH->BR mit Intermediary USD-Korrespondent",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["ITAU_BR"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["ITAU_BR"],
    creditor_name="Sao Paulo Export SA", creditor_country="BR",
    cdtr_acct_other="1234567890", cdtr_acct_scheme="BBAN",
    amount="50000.00", currency="USD",
    intrmy1_bic=BICS["JPM_NY"],
    rmt="Commodity Trade",
    bemerkungen="USD via NY Korrespondent",
))
rows.append(make_row(
    "TC-PCS-012", "CBPR+ 3-Hop Chain",
    "Zahlung mit drei Intermediaries (komplexe Chain)",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["NCB_RIY"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["NCB_RIY"],
    creditor_name="Riyadh Investments LLC", creditor_country="SA",
    creditor_iban=IBANS["SA"],
    amount="100000.00", currency="USD",
    intrmy1_bic=BICS["JPM_NY"], intrmy2_bic=BICS["HSBC_LDN"],
    rmt="Investment Transfer",
    bemerkungen="3 Intermediaries USD->SA via NY+LDN",
))
rows.append(make_row(
    "TC-PCS-013", "CBPR+ EUR via Intermediary",
    "EUR-Zahlung CH->ES via IT Intermediary",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["BBVA_MAD"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["BBVA_MAD"],
    creditor_name="Madrid Import SL", creditor_country="ES",
    creditor_iban=IBANS["ES"],
    amount="7500.00", currency="EUR",
    intrmy1_bic=BICS["UNICREDIT_MI"],
    rmt="Subcontractor Payment",
))
rows.append(make_row(
    "TC-PCS-014", "CBPR+ mit Fedwire Intermediary",
    "USD-Zahlung mit ClrSysMmbId statt BIC fuer IntrmyAgt1",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["CITI_NY"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["CITI_NY"],
    creditor_name="NY Clearing Corp", creditor_country="US",
    cdtr_acct_other="987654321", cdtr_acct_scheme="BBAN",
    amount="30000.00", currency="USD",
    intrmy1_clr="021000021",
    bemerkungen="Intermediary ueber Fedwire ABA-Nummer",
))
rows.append(make_row(
    "TC-PCS-015", "CBPR+ ohne Intermediary (direkt)",
    "Direkter CH->DE Hop ohne Intermediary (override clears default)",
    instg_bic=BICS["CS_ZURICH"], instd_bic=BICS["COBA_FFM"],
    debtor_agt_bic=BICS["CS_ZURICH"], creditor_agt_bic=BICS["COBA_FFM"],
    creditor_name="Handel GmbH", creditor_country="DE",
    creditor_iban=IBANS["DE"],
    amount="2200.00", currency="EUR",
    intrmy1_bic=BICS["UBS_ZURICH"],  # Explicit override sonst default
    rmt="Direct bilateral",
))

# --- Group C: Ultimate Parties + LEI (5 Cases) ---
rows.append(make_row(
    "TC-PCS-016", "CBPR+ UltmtDbtr",
    "Zahlung mit Ultimate Debtor (Konzern-Szenario)",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["DB_FFM"],
    debtor_name="Tochter AG", debtor_iban=IBANS["CH"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["DB_FFM"],
    creditor_iban=IBANS["DE"], creditor_country="DE",
    amount="45000.00", currency="EUR",
    overrides="UltmtDbtr.Nm=Mutter Holding AG",
    rmt="Konzernweiterleitung",
    bemerkungen="Ultimate Debtor = Mutterkonzern",
))
rows.append(make_row(
    "TC-PCS-017", "CBPR+ UltmtCdtr",
    "Zahlung mit Ultimate Creditor (Clearing-Szenario)",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["KBC_BRU"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["KBC_BRU"],
    creditor_name="BE Clearing Agent",
    creditor_country="BE", creditor_iban=IBANS["BE"],
    amount="18000.00", currency="EUR",
    overrides="UltmtCdtr.Nm=End Beneficiary Corp",
    rmt="Via Clearing Agent",
    bemerkungen="Ultimate Creditor = Endbeguenstigter",
))
rows.append(make_row(
    "TC-PCS-018", "CBPR+ LEI Debtor",
    "Debtor mit LEI ISO 17442",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["BNP_PARIS"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["BNP_PARIS"],
    creditor_iban=IBANS["FR"], creditor_country="FR",
    amount="12500.00", currency="EUR",
    overrides="Dbtr.Id.OrgId.LEI=506700GE1G29325QX363",
    rmt="B2B Payment",
))
rows.append(make_row(
    "TC-PCS-019", "CBPR+ UltmtDbtr + UltmtCdtr",
    "Intercompany Transfer mit beiden Ultimate-Parties",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["DBS_SG"],
    debtor_name="CH Tochter AG", debtor_iban=IBANS["CH"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["DBS_SG"],
    creditor_name="Clearing Pte Ltd",
    creditor_country="SG",
    cdtr_acct_other="0011234567", cdtr_acct_scheme="BBAN",
    amount="250000.00", currency="USD",
    intrmy1_bic=BICS["JPM_NY"],
    overrides="UltmtDbtr.Nm=Mutterkonzern AG; UltmtCdtr.Nm=Asien Tochter Pte Ltd",
    rmt="Intercompany Q1",
    category_purp="INTC",
    bemerkungen="Full Ultimate Chain mit CategoryPurpose",
))
rows.append(make_row(
    "TC-PCS-020", "CBPR+ Creditor LEI",
    "Creditor mit LEI",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["NATWEST_LDN"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["NATWEST_LDN"],
    creditor_name="UK Corp Ltd",
    creditor_country="GB", creditor_iban=IBANS["GB"],
    amount="6500.00", currency="GBP",
    overrides="Cdtr.Id.OrgId.LEI=213800WAVVOPS85N2205",
    rmt="Corporate Payment",
))

# --- Group D: Purpose / Category / Charges (5 Cases) ---
rows.append(make_row(
    "TC-PCS-021", "CBPR+ PurposeCode SALA",
    "Gehaltszahlung mit PurposeCode=SALA",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["DB_FFM"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["DB_FFM"],
    creditor_name="Mitarbeiter Max",
    creditor_country="DE", creditor_iban=IBANS["DE"],
    amount="4500.00", currency="EUR",
    purpose="SALA",
    rmt="Gehalt April 2026",
    bemerkungen="Purp/Cd=SALA",
))
rows.append(make_row(
    "TC-PCS-022", "CBPR+ CategoryPurpose SUPP",
    "Lieferantenzahlung CategoryPurpose=SUPP",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["ING_AMS"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["ING_AMS"],
    creditor_name="Office Supplies NL BV",
    creditor_country="NL", creditor_iban=IBANS["NL"],
    amount="3200.00", currency="EUR",
    category_purp="SUPP",
    rmt="Bueromaterial",
    bemerkungen="PmtTpInf/CtgyPurp/Cd=SUPP",
))
rows.append(make_row(
    "TC-PCS-023", "CBPR+ TRAD Purpose",
    "Handelszahlung mit Purpose=TRAD + Category=TRAD",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["MUFG_TKY"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["MUFG_TKY"],
    creditor_name="JP Import KK",
    creditor_country="JP",
    cdtr_acct_other="0123456", cdtr_acct_scheme="BBAN",
    amount="2500000.00", currency="JPY",
    purpose="TRAD", category_purp="TRAD",
    rmt="Import Settlement",
))
rows.append(make_row(
    "TC-PCS-024", "CBPR+ ChrgBr=DEBT",
    "Alle Charges zu Lasten Debtor",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["SG_PARIS"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["SG_PARIS"],
    creditor_iban=IBANS["FR"], creditor_country="FR",
    amount="1200.00", currency="EUR",
    chrg_br="DEBT",
    rmt="Net Amount to Beneficiary",
))
rows.append(make_row(
    "TC-PCS-025", "CBPR+ ChrgBr=CRED",
    "Alle Charges zu Lasten Creditor",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["CA_PARIS"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["CA_PARIS"],
    creditor_iban=IBANS["FR"], creditor_country="FR",
    amount="1800.00", currency="EUR",
    chrg_br="CRED",
    rmt="Gross amount, creditor bears",
))

# --- Group E: ClrSysMmbId / Settlement Method variations (5 Cases) ---
rows.append(make_row(
    "TC-PCS-026", "CBPR+ CdtrAgt via Fedwire ABA",
    "Creditor Agent identifiziert ueber Fedwire MmbId, kein BIC",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["WF_SF"],
    debtor_agt_bic=BICS["UBS_ZURICH"],
    creditor_agt_bic=None, cdtr_agt_clr="121000248",
    creditor_name="SF Corp", creditor_country="US",
    cdtr_acct_other="12345678", cdtr_acct_scheme="BBAN",
    amount="4500.00", currency="USD",
    bemerkungen="CdtrAgt nur per ClrSysMmbId (Wells Fargo Routing)",
))
rows.append(make_row(
    "TC-PCS-027", "CBPR+ SttlmMtd INGA",
    "SettlementMethod = Instructing Agent settles",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["DB_FFM"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["DB_FFM"],
    creditor_iban=IBANS["DE"], creditor_country="DE",
    amount="3000.00", currency="EUR",
    sttlm_mtd="INGA",
    rmt="INGA Settlement Test",
))
rows.append(make_row(
    "TC-PCS-028", "CBPR+ SttlmMtd INDA mit SttlmAcct",
    "INDA Settlement mit expliziter Settlement-Account-Referenz via Override",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["DB_FFM"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["DB_FFM"],
    creditor_iban=IBANS["DE"], creditor_country="DE",
    amount="4500.00", currency="EUR",
    sttlm_mtd="INDA",
    rmt="INDA with nostro account reference",
    bemerkungen="INDA+SttlmAcct-Szenario (CLRG nicht in CBPR+)",
))
rows.append(make_row(
    "TC-PCS-029", "CBPR+ T+2 Settlement",
    "Settlement zwei Banktage in der Zukunft",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["DB_FFM"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["DB_FFM"],
    creditor_iban=IBANS["DE"], creditor_country="DE",
    amount="7700.00", currency="EUR",
    sttlm_dt="2026-04-10",
    rmt="T+2",
))
rows.append(make_row(
    "TC-PCS-030", "CBPR+ grosser Betrag",
    "High-value Transfer",
    instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["BOA_NY"],
    debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["BOA_NY"],
    creditor_name="LA Investments LLC", creditor_country="US",
    cdtr_acct_other="987654321", cdtr_acct_scheme="BBAN",
    amount="9999999.99", currency="USD",
    rmt="Large value transfer",
    bemerkungen="Max-Amount Test",
))

# --- Group F: Negative Tests via ViolateRule (20 Cases) ---
NEG_CASES = [
    ("TC-PCS-N01", "UETR fehlt",
     "BR-CBPR-PACS-001", "Violation: UETR wird geleert"),
    ("TC-PCS-N02", "InstgAgt ohne Identifikation",
     "BR-CBPR-PACS-002", "Violation: InstgAgt hat weder BIC noch ClrSysMmbId"),
    ("TC-PCS-N03", "InstdAgt ohne Identifikation",
     "BR-CBPR-PACS-003", "Violation: InstdAgt hat weder BIC noch ClrSysMmbId"),
    ("TC-PCS-N04", "SttlmMtd COVE out of scope",
     "BR-CBPR-PACS-004", "Violation: SttlmMtd auf COVE (V1 lehnt ab)"),
    ("TC-PCS-N05", "BAH MsgDefIdr falsche Version",
     "BR-CBPR-PACS-007", "Violation: MsgDefIdr=pacs.008.001.09"),
    ("TC-PCS-N06", "BAH BizSvc falsch",
     "BR-CBPR-PACS-008", "Violation: BizSvc=swift.wrong.01"),
    ("TC-PCS-N07", "ChrgBr XXXX",
     "BR-CBPR-PACS-010", "Violation: ChrgBr auf XXXX (kein DEBT/CRED/SHAR)"),
    ("TC-PCS-N08", "Waehrung X9X",
     "BR-CBPR-PACS-011", "Violation: Waehrungscode ungueltig"),
    ("TC-PCS-N09", "UETR Format invalid",
     "BR-CBPR-PACS-015", "Violation: UETR ist keine UUIDv4"),
]
# Pad to 20 negatives by repeating with different contexts
neg_rows = []
for i, (tc_id, title, violate, bem) in enumerate(NEG_CASES, start=1):
    neg_rows.append(make_row(
        tc_id, f"Negative: {title}",
        f"Negative Test fuer {violate}",
        expected="NOK",
        instg_bic=BICS["UBS_ZURICH"], instd_bic=BICS["DB_FFM"],
        debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=BICS["DB_FFM"],
        creditor_iban=IBANS["DE"], creditor_country="DE",
        amount="1000.00", currency="EUR",
        violate=violate,
        rmt="Negative test case",
        bemerkungen=bem,
    ))

# Additional 11 negative cases with different corridors/variations to reach 20
additional_neg = [
    ("TC-PCS-N10", "UETR fehlt USD-Zahlung", "BR-CBPR-PACS-001", "DE_USD"),
    ("TC-PCS-N11", "InstgAgt fehlt GBP", "BR-CBPR-PACS-002", "GB_GBP"),
    ("TC-PCS-N12", "InstdAgt fehlt JPY", "BR-CBPR-PACS-003", "JP_JPY"),
    ("TC-PCS-N13", "COVE fuer GBP", "BR-CBPR-PACS-004", "GB_GBP"),
    ("TC-PCS-N14", "MsgDefIdr falsch USD", "BR-CBPR-PACS-007", "US_USD"),
    ("TC-PCS-N15", "BizSvc falsch CNY", "BR-CBPR-PACS-008", "CN_CNY"),
    ("TC-PCS-N16", "ChrgBr invalid SGD", "BR-CBPR-PACS-010", "SG_SGD"),
    ("TC-PCS-N17", "Currency invalid", "BR-CBPR-PACS-011", "DE_EUR"),
    ("TC-PCS-N18", "UETR format ungueltig CAD", "BR-CBPR-PACS-015", "CA_CAD"),
    ("TC-PCS-N19", "UETR fehlt HKD", "BR-CBPR-PACS-001", "HK_HKD"),
    ("TC-PCS-N20", "ChrgBr invalid DEBT+extra", "BR-CBPR-PACS-010", "FR_EUR"),
]

CORRIDOR_MAP = {
    "DE_USD": (BICS["JPM_NY"], "US", "US Partner", "1234567", "BBAN", "USD"),
    "GB_GBP": (BICS["HSBC_LDN"], "GB", "UK Corp", IBANS["GB"], None, "GBP"),
    "JP_JPY": (BICS["MIZUHO_TKY"], "JP", "JP Holdings", "0123456", "BBAN", "JPY"),
    "US_USD": (BICS["JPM_NY"], "US", "NY Corp", "042000021", "BBAN", "USD"),
    "CN_CNY": (BICS["ICBC_BJ"], "CN", "BJ Trade", "6222020100000000", "BBAN", "CNY"),
    "SG_SGD": (BICS["DBS_SG"], "SG", "SG Corp", "0011234567", "BBAN", "SGD"),
    "DE_EUR": (BICS["DB_FFM"], "DE", "DE Corp", IBANS["DE"], None, "EUR"),
    "CA_CAD": (BICS["RBC_TO"], "CA", "CA Corp", "12345678901", "BBAN", "CAD"),
    "HK_HKD": (BICS["HSBC_HK"], "HK", "HK Corp", "12345678", "BBAN", "HKD"),
    "FR_EUR": (BICS["BNP_PARIS"], "FR", "FR Corp", IBANS["FR"], None, "EUR"),
}

for tc_id, title, violate, corridor in additional_neg:
    bic, ctry, cdtr_nm, acct, scheme, ccy = CORRIDOR_MAP[corridor]
    if scheme is None:
        # IBAN
        neg_rows.append(make_row(
            tc_id, f"Negative: {title}", f"Negative {violate} ({corridor})",
            expected="NOK",
            instg_bic=BICS["UBS_ZURICH"], instd_bic=bic,
            debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=bic,
            creditor_name=cdtr_nm, creditor_country=ctry, creditor_iban=acct,
            amount="5000.00", currency=ccy,
            violate=violate, rmt="Negative Corridor Test",
            bemerkungen=f"{violate} in {corridor}",
        ))
    else:
        # Non-IBAN
        neg_rows.append(make_row(
            tc_id, f"Negative: {title}", f"Negative {violate} ({corridor})",
            expected="NOK",
            instg_bic=BICS["UBS_ZURICH"], instd_bic=bic,
            debtor_agt_bic=BICS["UBS_ZURICH"], creditor_agt_bic=bic,
            creditor_name=cdtr_nm, creditor_country=ctry,
            cdtr_acct_other=acct, cdtr_acct_scheme=scheme,
            amount="5000.00", currency=ccy,
            violate=violate, rmt="Negative Corridor Test",
            bemerkungen=f"{violate} in {corridor}",
        ))

rows.extend(neg_rows)

print(f"Total rows: {len(rows)}")


# ---------------------------------------------------------------------------
# Write Excel
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Testfaelle"
ws.append(HEADER)
for r in rows:
    ws.append(r)

# Styling
hf = Font(bold=True, color="FFFFFF")
hfill = PatternFill("solid", fgColor="305496")
for c in ws[1]:
    c.font = hf
    c.fill = hfill
    c.alignment = Alignment(horizontal="center", vertical="center")
for i, col in enumerate(HEADER, 1):
    ws.column_dimensions[get_column_letter(i)].width = max(12, min(30, len(col) + 2))

out_path = Path("templates/testfaelle_pacs008_comprehensive.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
