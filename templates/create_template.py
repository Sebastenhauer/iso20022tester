"""Erstellt die Beispiel-Excel-Vorlage mit 10 Testfällen."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Testfälle"

# Header
headers = [
    "TestcaseID", "Titel", "Ziel", "Erwartetes Ergebnis", "Zahlungstyp",
    "Betrag", "Währung", "Debtor Infos", "Weitere Testdaten",
    "Erwartete API-Antwort", "Ergebnis (OK/NOK)", "Bemerkungen",
]

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font

# Debtor-Info (gleich für alle Testfälle)
debtor = "Name=Muster AG; IBAN=CH5604835012345678009; BIC=CRESCHZZ80A; Strasse=Bahnhofstrasse; Hausnummer=1; PLZ=8001; Ort=Zuerich; Land=CH"

# Testfälle
testcases = [
    # --- SEPA ---
    {
        "id": "TC-SEPA-001", "titel": "SEPA Standardzahlung",
        "ziel": "Positive SEPA EUR-Zahlung generieren",
        "ergebnis": "OK", "typ": "SEPA", "betrag": 1500.00, "waehrung": "EUR",
        "debtor": debtor, "overrides": "", "bemerkung": "",
    },
    {
        "id": "TC-SEPA-002", "titel": "SEPA falsche Waehrung",
        "ziel": "Negative SEPA-Zahlung: Waehrung wird auf CHF gesetzt",
        "ergebnis": "NOK", "typ": "SEPA", "betrag": 250.50, "waehrung": "EUR",
        "debtor": debtor, "overrides": "ViolateRule=BR-SEPA-001",
        "bemerkung": "Erwartet: BR-SEPA-001 verletzt (Waehrung != EUR)",
    },
    {
        "id": "TC-SEPA-003", "titel": "SEPA mit Creditor-Override",
        "ziel": "SEPA-Zahlung mit explizitem Creditor-Namen und BIC",
        "ergebnis": "OK", "typ": "SEPA", "betrag": 4999.99, "waehrung": "EUR",
        "debtor": debtor,
        "overrides": "Cdtr.Nm=Beispiel GmbH; CdtrAgt.BICFI=COBADEFFXXX",
        "bemerkung": "Creditor wird explizit gesetzt statt generiert",
    },
    # --- Domestic-QR ---
    {
        "id": "TC-QR-001", "titel": "QR-Zahlung Standard CHF",
        "ziel": "Positive QR-Zahlung mit QR-IBAN und QRR",
        "ergebnis": "OK", "typ": "Domestic-QR", "betrag": 3200.00, "waehrung": "CHF",
        "debtor": debtor, "overrides": "", "bemerkung": "",
    },
    {
        "id": "TC-QR-002", "titel": "QR-Zahlung ohne Referenz",
        "ziel": "Negative QR-Zahlung: QRR-Referenz wird entfernt",
        "ergebnis": "NOK", "typ": "Domestic-QR", "betrag": 890.00, "waehrung": "CHF",
        "debtor": debtor, "overrides": "ViolateRule=BR-QR-002",
        "bemerkung": "Erwartet: BR-QR-002 verletzt (keine QRR bei QR-IBAN)",
    },
    # --- Domestic-IBAN ---
    {
        "id": "TC-IBAN-001", "titel": "Domestic IBAN Standard",
        "ziel": "Positive Inlandszahlung mit regulaerer CH-IBAN",
        "ergebnis": "OK", "typ": "Domestic-IBAN", "betrag": 5000.00, "waehrung": "CHF",
        "debtor": debtor, "overrides": "", "bemerkung": "",
    },
    {
        "id": "TC-IBAN-002", "titel": "Domestic IBAN falsche Waehrung",
        "ziel": "Negative Inlandszahlung: Waehrung wird auf EUR gesetzt",
        "ergebnis": "NOK", "typ": "Domestic-IBAN", "betrag": 750.00, "waehrung": "CHF",
        "debtor": debtor, "overrides": "ViolateRule=BR-IBAN-004",
        "bemerkung": "Erwartet: BR-IBAN-004 verletzt (Waehrung != CHF)",
    },
    # --- CBPR+ ---
    {
        "id": "TC-CBPR-001", "titel": "CBPR+ USD-Zahlung",
        "ziel": "Positive Cross-Border-Zahlung in USD",
        "ergebnis": "OK", "typ": "CBPR+", "betrag": 10000.00, "waehrung": "USD",
        "debtor": debtor, "overrides": "CdtrAgt.BICFI=BNPAFRPP",
        "bemerkung": "",
    },
    {
        "id": "TC-CBPR-002", "titel": "CBPR+ ohne Agent",
        "ziel": "Negative CBPR+: Creditor-Agent BIC wird entfernt",
        "ergebnis": "NOK", "typ": "CBPR+", "betrag": 2500.00, "waehrung": "GBP",
        "debtor": debtor, "overrides": "CdtrAgt.BICFI=BARCGB22; ViolateRule=BR-CBPR-005",
        "bemerkung": "Erwartet: BR-CBPR-005 verletzt (kein Creditor-Agent)",
    },
    {
        "id": "TC-CBPR-003", "titel": "CBPR+ JPY-Zahlung",
        "ziel": "Positive Cross-Border-Zahlung in JPY",
        "ergebnis": "OK", "typ": "CBPR+", "betrag": 500000.00, "waehrung": "JPY",
        "debtor": debtor, "overrides": "CdtrAgt.BICFI=BOABORJ1",
        "bemerkung": "Japanische Yen-Zahlung",
    },
]

for i, tc in enumerate(testcases, 2):
    ws.cell(row=i, column=1, value=tc["id"])
    ws.cell(row=i, column=2, value=tc["titel"])
    ws.cell(row=i, column=3, value=tc["ziel"])
    ws.cell(row=i, column=4, value=tc["ergebnis"])
    ws.cell(row=i, column=5, value=tc["typ"])
    ws.cell(row=i, column=6, value=tc["betrag"])
    ws.cell(row=i, column=7, value=tc["waehrung"])
    ws.cell(row=i, column=8, value=tc["debtor"])
    ws.cell(row=i, column=9, value=tc["overrides"])
    ws.cell(row=i, column=10, value="")
    ws.cell(row=i, column=11, value="")
    ws.cell(row=i, column=12, value=tc["bemerkung"])

# Spaltenbreiten
widths = [15, 30, 40, 18, 16, 10, 10, 80, 40, 20, 15, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

wb.save("templates/testfaelle_vorlage.xlsx")
print("Template erstellt: templates/testfaelle_vorlage.xlsx")
