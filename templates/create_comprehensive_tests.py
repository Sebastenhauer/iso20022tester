"""Erstellt umfassende Testfaelle: Top-30 Waehrungen/Laender, Charges, RmtInf."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Testfaelle"

headers = [
    "TestcaseID", "Titel", "Ziel", "Erwartetes Ergebnis", "Zahlungstyp",
    "Debtor Name", "Debtor IBAN", "Debtor BIC",
    "Betrag", "Waehrung",
    "Creditor Name", "Creditor IBAN", "Creditor BIC",
    "Verwendungszweck", "ViolateRule", "Weitere Testdaten",
    "Erwartete API-Antwort", "Ergebnis (OK/NOK)", "Bemerkungen",
]

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font

DEBTOR_IBAN = "CH5604835012345678009"

# BICs fuer Top-30 Laender (repraesentativ)
COUNTRY_BICS = {
    "US": "CHASUS33XXX", "GB": "BARCGB22XXX", "DE": "DEUTDEFFXXX",
    "FR": "BNPAFRPPXXX", "JP": "BOABORJ1XXX", "CN": "BKCHCNBJXXX",
    "CA": "ROYCCAT2XXX", "AU": "ANZBAU3MXXX", "IN": "SBININBBXXX",
    "BR": "BRAABORJXXX", "KR": "KOABORJ1XXX", "MX": "BCMRMXMMXXX",
    "SG": "DBSSSGSGXXX", "HK": "HSBCHKHHXXX", "SE": "SWEDSESSXXX",
    "NO": "DNBANOKK",    "DK": "DABADKKKXXX", "ZA": "SBZAZAJJXXX",
    "SA": "SABBSARX",    "AE": "NBABORJXXXX", "TH": "BKKBTHBKXXX",
    "TR": "ISBKTRISXXX", "PL": "BPKOPLPWXXX", "NL": "ABNANL2AXXX",
    "BE": "GEBABEBB",    "AT": "BKAUATWWXXX", "IT": "BCITITMM",
    "ES": "BBVAESMMXXX", "PT": "CGDIPTPL",    "FI": "NDEAFIHH",
}

# Top-30 Waehrungen mit Laenderzuordnung
CURRENCIES = [
    ("USD", "US", "US-Dollar"),
    ("EUR", "DE", "Euro"),
    ("JPY", "JP", "Japanischer Yen"),
    ("GBP", "GB", "Britisches Pfund"),
    ("CNY", "CN", "Chinesischer Yuan"),
    ("AUD", "AU", "Australischer Dollar"),
    ("CAD", "CA", "Kanadischer Dollar"),
    ("CHF", "CH", "Schweizer Franken"),
    ("HKD", "HK", "Hongkong-Dollar"),
    ("SGD", "SG", "Singapur-Dollar"),
    ("SEK", "SE", "Schwedische Krone"),
    ("KRW", "KR", "Suedkoreanischer Won"),
    ("NOK", "NO", "Norwegische Krone"),
    ("NZD", "AU", "Neuseeland-Dollar"),  # AU als Proxy
    ("INR", "IN", "Indische Rupie"),
    ("MXN", "MX", "Mexikanischer Peso"),
    ("TWD", "HK", "Taiwan-Dollar"),  # HK als Proxy
    ("ZAR", "ZA", "Suedafrikanischer Rand"),
    ("BRL", "BR", "Brasilianischer Real"),
    ("DKK", "DK", "Daenische Krone"),
    ("PLN", "PL", "Polnischer Zloty"),
    ("THB", "TH", "Thailaendischer Baht"),
    ("ILS", "IL", "Israelischer Schekel"),
    ("IDR", "SG", "Indonesische Rupiah"),  # SG als Proxy
    ("CZK", "CZ", "Tschechische Krone"),
    ("AED", "AE", "VAE-Dirham"),
    ("TRY", "TR", "Tuerkische Lira"),
    ("HUF", "HU", "Ungarischer Forint"),
    ("SAR", "SA", "Saudi-Riyal"),
    ("PHP", "SG", "Philippinischer Peso"),  # SG als Proxy
]

# Charge Bearer Codes fuer CBPR+
CHARGE_BEARERS = ["DEBT", "CRED", "SHAR"]

row_num = 2
tc_counter = 0


def add_row(tc_id, titel, ziel, ergebnis, typ, betrag, waehrung,
            cdtr_name=None, cdtr_bic=None, violate=None, overrides=None,
            bemerkung=None):
    global row_num
    ws.cell(row=row_num, column=1, value=tc_id)
    ws.cell(row=row_num, column=2, value=titel)
    ws.cell(row=row_num, column=3, value=ziel)
    ws.cell(row=row_num, column=4, value=ergebnis)
    ws.cell(row=row_num, column=5, value=typ)
    ws.cell(row=row_num, column=7, value=DEBTOR_IBAN)
    ws.cell(row=row_num, column=9, value=betrag)
    ws.cell(row=row_num, column=10, value=waehrung)
    if cdtr_name:
        ws.cell(row=row_num, column=11, value=cdtr_name)
    if cdtr_bic:
        ws.cell(row=row_num, column=13, value=cdtr_bic)
    if violate:
        ws.cell(row=row_num, column=15, value=violate)
    if overrides:
        ws.cell(row=row_num, column=16, value=overrides)
    if bemerkung:
        ws.cell(row=row_num, column=19, value=bemerkung)
    row_num += 1


# ============================================================
# SEPA Basis-Tests (EUR only)
# ============================================================

add_row("TC-S-001", "SEPA Standard EUR", "Positive SEPA EUR", "OK",
        "SEPA", 1500.00, "EUR", bemerkung="Basis-SEPA")
add_row("TC-S-002", "SEPA falsche Waehrung", "SEPA mit CHF statt EUR", "NOK",
        "SEPA", 250.50, "EUR", violate="BR-SEPA-001")
add_row("TC-S-003", "SEPA falscher ChrgBr", "SEPA ChrgBr != SLEV", "NOK",
        "SEPA", 500.00, "EUR", violate="BR-SEPA-003")
add_row("TC-S-004", "SEPA Name zu lang", "Creditor > 70 Zeichen", "NOK",
        "SEPA", 100.00, "EUR", violate="BR-SEPA-004")

# ============================================================
# Domestic QR Tests
# ============================================================

add_row("TC-QR-001", "QR Standard CHF", "QR-Zahlung CHF", "OK",
        "Domestic-QR", 3200.00, "CHF")
add_row("TC-QR-002", "QR ohne Referenz", "QRR entfernt", "NOK",
        "Domestic-QR", 890.00, "CHF", violate="BR-QR-002")
add_row("TC-QR-003", "QR mit SCOR", "SCOR statt QRR", "NOK",
        "Domestic-QR", 500.00, "CHF", violate="BR-QR-003")
add_row("TC-QR-004", "QR falsche Waehrung", "USD statt CHF/EUR", "NOK",
        "Domestic-QR", 1000.00, "CHF", violate="BR-QR-004")

# ============================================================
# Domestic IBAN Tests
# ============================================================

add_row("TC-DI-001", "Domestic IBAN Standard", "Inlandszahlung CHF", "OK",
        "Domestic-IBAN", 5000.00, "CHF")
add_row("TC-DI-002", "Domestic IBAN QR-IBAN", "QR-IBAN als Creditor", "NOK",
        "Domestic-IBAN", 750.00, "CHF", violate="BR-IBAN-001")
add_row("TC-DI-003", "Domestic IBAN mit QRR", "QRR bei regulaerer IBAN", "NOK",
        "Domestic-IBAN", 1200.00, "CHF", violate="BR-IBAN-002")
add_row("TC-DI-004", "Domestic IBAN falsche Waehrung", "EUR statt CHF", "NOK",
        "Domestic-IBAN", 800.00, "CHF", violate="BR-IBAN-004")

# ============================================================
# CBPR+ Charge Bearer Tests (OUR, SHA, BEN/CRED)
# ============================================================

for cb in CHARGE_BEARERS:
    add_row(
        f"TC-CB-{cb}", f"CBPR+ ChrgBr {cb}",
        f"CBPR+ mit ChrgBr={cb}", "OK",
        "CBPR+", 10000.00, "USD",
        cdtr_bic="CHASUS33XXX",
        overrides=f"ChrgBr={cb}",
        bemerkung=f"Charge Bearer {cb}",
    )

# CBPR+ mit ungueltigem ChrgBr
add_row("TC-CB-INV", "CBPR+ ungueltiger ChrgBr", "ChrgBr INVALID", "NOK",
        "CBPR+", 5000.00, "USD", cdtr_bic="CHASUS33XXX",
        violate="BR-CBPR-003", overrides="ChrgBr=SHAR",
        bemerkung="Violation setzt ChrgBr auf INVALID")

# CBPR+ ohne Agent
add_row("TC-CB-NOAGT", "CBPR+ ohne Agent", "Fehlender CdtrAgt", "NOK",
        "CBPR+", 2500.00, "GBP", cdtr_bic="BARCGB22XXX",
        violate="BR-CBPR-005")

# ============================================================
# Remittance Information Tests
# ============================================================

add_row("TC-RMT-001", "CBPR+ mit Verwendungszweck", "USTRD Remittance", "OK",
        "CBPR+", 7500.00, "USD", cdtr_bic="CHASUS33XXX",
        bemerkung="Unstrukturierte Remittance Info")
# Zeile hat Verwendungszweck in Spalte N
ws.cell(row=row_num - 1, column=14, value="Invoice INV-2026-042 Payment for services")

add_row("TC-RMT-002", "USTRD max 140 Zeichen", "USTRD genau 140 Zeichen (grenzwert)", "OK",
        "CBPR+", 3000.00, "EUR", cdtr_bic="DEUTDEFFXXX",
        bemerkung="Grenzwert-Test: genau 140 Zeichen")
# Setze 140-Zeichen-Verwendungszweck
ws.cell(row=row_num - 1, column=14, value="A" * 140)

# ============================================================
# Currency Code Tests
# ============================================================

add_row("TC-CCY-OK", "Exotische Waehrung KWD", "Waehrungscode KWD (Kuwait-Dinar)", "OK",
        "CBPR+", 1000.00, "KWD", cdtr_bic="CHASUS33XXX",
        bemerkung="Exotische aber gueltige Waehrung")

# ============================================================
# Top-30 Waehrungen: CBPR+ Happy Path
# ============================================================

for ccy, country, name in CURRENCIES:
    if ccy == "CHF":
        continue  # CHF ist Domestic, nicht CBPR+
    if ccy == "EUR":
        continue  # EUR ist SEPA

    bic = COUNTRY_BICS.get(country, "CHASUS33XXX")

    add_row(
        f"TC-CCY-{ccy}", f"CBPR+ {ccy} ({name})",
        f"Positive CBPR+ Zahlung in {ccy}", "OK",
        "CBPR+", 10000.00, ccy,
        cdtr_bic=bic,
        bemerkung=f"Waehrung {ccy}, Land {country}",
    )

# ============================================================
# Top-30 Waehrungen: CBPR+ Negative (fehlender Agent)
# ============================================================

for ccy, country, name in CURRENCIES:
    if ccy in ("CHF", "EUR"):
        continue

    bic = COUNTRY_BICS.get(country, "CHASUS33XXX")

    add_row(
        f"TC-CCY-{ccy}-NOK", f"CBPR+ {ccy} ohne Agent",
        f"Negative CBPR+ {ccy}: Agent fehlt", "NOK",
        "CBPR+", 5000.00, ccy,
        cdtr_bic=bic,
        violate="BR-CBPR-005",
        bemerkung=f"Agent wird entfernt fuer {ccy}",
    )

# ============================================================
# CBPR+ pro SEPA-Land mit verschiedenen Waehrungen (nicht-EUR)
# ============================================================

sepa_countries_sample = [
    ("DE", "DEUTDEFFXXX", "USD"), ("FR", "BNPAFRPPXXX", "GBP"),
    ("IT", "BCITITMM", "JPY"), ("ES", "BBVAESMMXXX", "CHF"),
    ("NL", "ABNANL2AXXX", "USD"), ("BE", "GEBABEBB", "CAD"),
    ("AT", "BKAUATWWXXX", "AUD"), ("PT", "CGDIPTPL", "BRL"),
]

for country, bic, ccy in sepa_countries_sample:
    add_row(
        f"TC-XC-{country}-{ccy}", f"CBPR+ {country} in {ccy}",
        f"Cross-Border {country} in Fremdwaehrung {ccy}", "OK",
        "CBPR+", 25000.00, ccy,
        cdtr_bic=bic,
        bemerkung=f"SEPA-Land {country} mit Nicht-EUR Waehrung",
    )

# ============================================================
# Spaltenbreiten
# ============================================================

col_widths = {
    "A": 18, "B": 35, "C": 45, "D": 18, "E": 16,
    "F": 20, "G": 25, "H": 15, "I": 12, "J": 10,
    "K": 20, "L": 25, "M": 15, "N": 40,
    "O": 15, "P": 35, "Q": 20, "R": 15, "S": 40,
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

total_rows = row_num - 2
wb.save("templates/testfaelle_comprehensive.xlsx")
print(f"Comprehensive Test-Excel erstellt: {total_rows} Testfaelle")
print(f"  templates/testfaelle_comprehensive.xlsx")
